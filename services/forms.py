
from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

from config import OFFICIAL_INVENTORY_TEMPLATE_PATH

PROVENIENCIAS_PADRAO = [
    "Direção Geral",
    "Direção de Administração",
    "Direção de Ensino",
    "Direção de Pesquisa e Pós-Graduação",
    "Direção de Extensão",
    "Departamento",
    "Secretaria Acadêmica",
    "Biblioteca",
    "Coordenação de Curso",
    "Setor de Recursos Humanos",
    "Setor Financeiro",
    "Arquivo Setorial",
    "Outro",
]

LOOKUP_SHEET_NAME = "_lookup"
HELP_SHEET_NAME = "Ajuda de preenchimento"
MAIN_SHEET_NAME = "ANEXO II - Atividades Fim"

def _copy_row_style(ws, template_row: int, target_row: int, start_col: int = 2, end_col: int = 13) -> None:
    for col in range(start_col, end_col + 1):
        source = ws.cell(template_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.protection:
            target.protection = copy(source.protection)

def build_quick_fill_workbook(df_ttd: pd.DataFrame, max_rows: int = 300) -> bytes:
    wb = load_workbook(OFFICIAL_INVENTORY_TEMPLATE_PATH)
    ws = wb[wb.sheetnames[0]]

    # Replicate official blank row style through desired range
    for row_idx in range(11, 11 + max_rows):
        if row_idx > 11:
            _copy_row_style(ws, 11, row_idx)

    # Hidden lookup sheet
    if LOOKUP_SHEET_NAME in wb.sheetnames:
        del wb[LOOKUP_SHEET_NAME]
    lookup = wb.create_sheet(LOOKUP_SHEET_NAME)

    lookup_headers = [
        "codigo_classificacao", "item_documental", "natureza_documental", "grupo", "subgrupo",
        "serie", "subserie", "dossie_processo", "prazo_corrente", "prazo_corrente_anos",
        "prazo_intermediario", "prazo_intermediario_anos", "total_prazo_anos",
        "destinacao_final", "guarda_permanente"
    ]
    lookup.append(lookup_headers)

    cleaned = (
        df_ttd.sort_values(["source_priority", "codigo_classificacao", "item_documental"])
        .drop_duplicates(subset=["codigo_classificacao"], keep="first")
        .copy()
    )
    cleaned["guarda_permanente"] = cleaned["destinacao_final"].apply(
        lambda x: "Guarda permanente" if "permanente" in str(x).lower() else "-"
    )

    for _, row in cleaned.iterrows():
        lookup.append([
            row.get("codigo_classificacao", ""),
            row.get("item_documental", ""),
            row.get("natureza_documental", ""),
            row.get("grupo", ""),
            row.get("subgrupo", ""),
            row.get("serie", ""),
            row.get("subserie", ""),
            row.get("dossie_processo", ""),
            row.get("prazo_corrente", ""),
            row.get("prazo_corrente_anos", 0),
            row.get("prazo_intermediario", ""),
            row.get("prazo_intermediario_anos", 0),
            row.get("total_prazo_anos", 0),
            row.get("destinacao_final", ""),
            row.get("guarda_permanente", "-"),
        ])

    lookup.sheet_state = "hidden"

    # Help sheet
    if HELP_SHEET_NAME in wb.sheetnames:
        del wb[HELP_SHEET_NAME]
    help_ws = wb.create_sheet(HELP_SHEET_NAME)
    help_ws["A1"] = "Como preencher o inventário"
    help_ws["A1"].font = Font(bold=True, size=14)
    help_lines = [
        "1. Selecione o código na coluna 'Código Segundo o Plano de Classificação'.",
        "2. Preencha a proveniência, o ano de emissão, a referência, o assunto e o nº da caixa.",
        "3. Classe segundo a TTD, prazos, destinação, total e guarda permanente são preenchidos automaticamente.",
        "4. Se a destinação for guarda permanente, o campo 'Eliminação no ano de' mostrará '-'.",
        "5. O ano de eliminação é calculado como: ano de emissão + fase corrente + fase intermediária, quando houver eliminação.",
    ]
    for i, line in enumerate(help_lines, start=3):
        help_ws[f"A{i}"] = line
    help_ws.column_dimensions["A"].width = 120

    # Data validations
    code_dv = DataValidation(
        type="list",
        formula1=f"='{LOOKUP_SHEET_NAME}'!$A$2:$A${lookup.max_row}",
        allow_blank=True,
        showDropDown=True,
    )
    prov_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(PROVENIENCIAS_PADRAO)}"',
        allow_blank=True,
        showDropDown=True,
    )
    year_dv = DataValidation(type="whole", operator="between", formula1="1900", formula2="2100", allow_blank=True)

    ws.add_data_validation(code_dv)
    ws.add_data_validation(prov_dv)
    ws.add_data_validation(year_dv)

    for row_idx in range(11, 11 + max_rows):
        code_cell = f"B{row_idx}"
        prov_cell = f"C{row_idx}"
        year_cell = f"D{row_idx}"

        code_dv.add(ws[code_cell])
        prov_dv.add(ws[prov_cell])
        year_dv.add(ws[year_cell])

        ws[f"G{row_idx}"] = f'=IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$B:$B,""),"")'
        ws[f"H{row_idx}"] = f'=IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$J:$J,""),"")'
        ws[f"I{row_idx}"] = f'=IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$L:$L,""),"")'
        ws[f"J{row_idx}"] = f'=IF($B{row_idx}="","",IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$M:$M,""),""))'
        ws[f"K{row_idx}"] = (
            f'=IF(OR($B{row_idx}="",$D{row_idx}=""),"",'
            f'IF(IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$O:$O,""),"-")="Guarda permanente","-",'
            f'VALUE($D{row_idx})+IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$M:$M,0),0)))'
        )
        ws[f"L{row_idx}"] = f'=IFERROR(XLOOKUP($B{row_idx},{LOOKUP_SHEET_NAME}!$A:$A,{LOOKUP_SHEET_NAME}!$O:$O,""),"")'

    ws["B11"].comment = Comment("Escolha o código oficial de classificação.", "ChatGPT")
    ws["C11"].comment = Comment("Selecione a proveniência do documento.", "ChatGPT")
    ws["D11"].comment = Comment("Informe o ano de emissão. O ano de eliminação será calculado automaticamente quando houver eliminação.", "ChatGPT")
    ws["G11"].comment = Comment("Classe segundo a TTD preenchida automaticamente a partir do código.", "ChatGPT")
    ws.freeze_panes = "B11"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def parse_inventory_workbook(uploaded_file, df_ttd: pd.DataFrame) -> pd.DataFrame:
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    lookup = (
        df_ttd.sort_values(["source_priority", "codigo_classificacao", "item_documental"])
        .drop_duplicates(subset=["codigo_classificacao"], keep="first")
        .set_index("codigo_classificacao")
    )

    rows = []
    for row_idx in range(11, ws.max_row + 1):
        codigo = str(ws[f"B{row_idx}"].value or "").strip()
        proveniencia = str(ws[f"C{row_idx}"].value or "").strip()
        ano_emissao = ws[f"D{row_idx}"].value
        referencia = str(ws[f"E{row_idx}"].value or "").strip()
        assunto = str(ws[f"F{row_idx}"].value or "").strip()
        numero_caixa = str(ws[f"M{row_idx}"].value or "").strip()

        if not any([codigo, proveniencia, ano_emissao, referencia, assunto, numero_caixa]):
            continue

        row = lookup.loc[codigo] if codigo in lookup.index else None
        texto_obs = []
        if referencia:
            texto_obs.append(f"Referência: {referencia}")
        if assunto:
            texto_obs.append(f"Assunto: {assunto}")

        rows.append({
            "setor": proveniencia,
            "tipo_documental": row["item_documental"] if row is not None else assunto,
            "natureza_documental": row["natureza_documental"] if row is not None else "",
            "grupo": row["grupo"] if row is not None else "",
            "subgrupo": row["subgrupo"] if row is not None else "",
            "serie": row["serie"] if row is not None else "",
            "subserie": row["subserie"] if row is not None else "",
            "dossie_processo": row["dossie_processo"] if row is not None else "",
            "item_documental": row["item_documental"] if row is not None else assunto,
            "prazo_corrente": row["prazo_corrente"] if row is not None else "",
            "prazo_intermediario": row["prazo_intermediario"] if row is not None else "",
            "destinacao_final": row["destinacao_final"] if row is not None else "",
            "datas_limite": str(ano_emissao or ""),
            "quantidade": 1,
            "caixa": numero_caixa,
            "observacoes": " | ".join(texto_obs),
        })

    return pd.DataFrame(rows)
